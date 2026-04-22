"""XlsxRenderer — openpyxl + finance-colour-convention helpers.

SKILL convention:
  input     → blue font   (hardcoded assumptions)
  formula   → black font
  crossref  → green font
  external  → red font
  assumption→ yellow fill
  header    → white on palette[0]

Formulas are always strings like ``=SUM(B2:B10)`` — the planner supplies the
formula body without the leading ``=`` and we prepend it here.
"""

from __future__ import annotations

import re
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..ir import XlsxCell, XlsxIR, XlsxSheet
from .base import BaseRenderer, RenderError, RenderResult


_ROLE_FONT = {
    "input": "0042A3",      # blue  (hardcoded input)
    "formula": "000000",    # black
    "crossref": "1F8A38",   # green
    "external": "B91C1C",   # red
    "assumption": "000000", # black on yellow fill
    "label": "1F2937",
    "header": "FFFFFF",
}

_ROLE_FILL = {
    "assumption": "FDE68A",  # yellow
    "header": None,  # header takes palette[0]
}


_NUMBER_FORMATS = {
    "general": "General",
    "integer": "0",
    "number_2dp": "0.00",
    "currency_usd": "$#,##0.00",
    "percent": "0.00%",
    "date": "mm/dd/yyyy",
    "iso_date": "yyyy-mm-dd",
}


class XlsxRenderer(BaseRenderer):
    format = "xlsx"

    def _cell_font(self, c: XlsxCell, header_fill_hex: str | None) -> Font:
        hex_color = _ROLE_FONT.get(c.role, "000000")
        if c.font_color is not None:
            hex_color = c.font_color.value
        return Font(
            bold=c.bold or c.role == "header",
            italic=c.italic,
            color=hex_color,
        )

    def _cell_fill(self, c: XlsxCell, header_fill_hex: str | None) -> PatternFill | None:
        if c.role == "header" and header_fill_hex:
            return PatternFill("solid", fgColor=header_fill_hex)
        if c.fill is not None:
            return PatternFill("solid", fgColor=c.fill.value)
        if c.role == "assumption":
            return PatternFill("solid", fgColor=_ROLE_FILL["assumption"])
        return None

    def _apply_cell(self, ws_cell, c: XlsxCell, header_fill_hex: str | None) -> None:
        if c.formula:
            ws_cell.value = f"={c.formula}"
        else:
            ws_cell.value = c.value
        ws_cell.number_format = _NUMBER_FORMATS.get(c.number_format, "General")
        ws_cell.font = self._cell_font(c, header_fill_hex)
        fill = self._cell_fill(c, header_fill_hex)
        if fill is not None:
            ws_cell.fill = fill
        if c.align:
            ws_cell.alignment = Alignment(horizontal=c.align)
        # All cells get a subtle border for readability.
        thin = Side(border_style="thin", color="E5E7EB")
        ws_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_sheet(self, wb: Workbook, sheet: XlsxSheet, header_fill_hex: str | None) -> None:
        # Skip the default "Sheet" only for the first real sheet we add.
        ws = wb.create_sheet(title=sheet.name)

        # Column widths
        for col in sheet.columns:
            ws.column_dimensions[get_column_letter(col.index + 1)].width = col.width_chars

        for r_idx, row in enumerate(sheet.rows, start=1):
            for c_idx, cell in enumerate(row.cells, start=1):
                ws_cell = ws.cell(row=r_idx, column=c_idx)
                self._apply_cell(ws_cell, cell, header_fill_hex)

        if sheet.freeze_header_row and len(sheet.rows) >= 2:
            ws.freeze_panes = "A2"

    async def render(self, ir: XlsxIR, out_dir: Path) -> RenderResult:
        if not isinstance(ir, XlsxIR):
            raise RenderError(f"XlsxRenderer got wrong IR type: {type(ir).__name__}")
        started = time.perf_counter()
        wb = Workbook()
        default_sheet = wb.active
        # Remove the auto-created sheet; we'll make named ones.
        wb.remove(default_sheet)

        header_fill_hex = ir.theme.palette[0].value if ir.theme.palette else "0F172A"

        for sheet in ir.content.sheets:
            self._write_sheet(wb, sheet, header_fill_hex)

        out_dir.mkdir(parents=True, exist_ok=True)
        safe_name = re.sub(r"[^A-Za-z0-9_\-\.]", "_", ir.metadata.title)[:80] or "workbook"
        path = out_dir / f"{safe_name}.xlsx"
        wb.save(str(path))

        return RenderResult(
            path=path,
            doc_type="xlsx",
            bytes_size=path.stat().st_size,
            duration_ms=int((time.perf_counter() - started) * 1000),
            extra={"sheets": len(ir.content.sheets)},
        )

    async def fix(self, ir: XlsxIR, critic_findings, out_dir: Path) -> RenderResult:
        return await self.render(ir, out_dir)
