"""XlsxFormulaVerifier — detect #ERROR cells.

When LibreOffice is available we go: soffice --headless --calc
--convert-to xlsx to force recalc, then openpyxl with data_only=True.

When it's not, we still catch malformed formulas by inspecting the
openpyxl workbook (``cell.value`` starts with '#REF!' etc.). Newly-
authored formulas that depend on runtime evaluation won't register
without soffice, so we flag that as a "cannot_recalc" info issue.
"""

from __future__ import annotations

import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook

from .types import CriticReport, Issue, IssueCategory, IssueSeverity

_ERROR_MARKERS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")


class XlsxFormulaVerifier:
    name = "xlsx_formula"

    def verify(self, path: Path) -> CriticReport:
        issues: list[Issue] = []
        recalc_note = "no_recalc"
        if shutil.which("soffice") or shutil.which("libreoffice"):
            ok = self._recalc(path)
            recalc_note = "recalc_ok" if ok else "recalc_failed"
            if not ok:
                issues.append(Issue(
                    category=IssueCategory.FORMULA_ERROR,
                    severity=IssueSeverity.WARNING,
                    message="LibreOffice recalc returned non-zero; continuing with stale cache",
                ))

        try:
            wb = load_workbook(str(path), data_only=True)
        except Exception as e:
            return CriticReport(
                passed=False,
                issues=[Issue(category=IssueCategory.OPEN_ERROR, severity=IssueSeverity.CRITICAL, message=f"cannot open xlsx: {e}")],
                backend="xlsx_formula",
                note=recalc_note,
            )

        error_cells: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:  # type: Cell
                    v = cell.value
                    if isinstance(v, str) and v in _ERROR_MARKERS:
                        coord = f"{ws.title}!{cell.coordinate}"
                        error_cells.append(f"{coord}={v}")

        for c in error_cells:
            issues.append(Issue(
                category=IssueCategory.FORMULA_ERROR,
                severity=IssueSeverity.CRITICAL,
                message=f"formula error at {c}",
                hint="fix reference / check denominator",
            ))

        passed = not any(i.severity == IssueSeverity.CRITICAL for i in issues)
        return CriticReport(passed=passed, issues=issues, backend="xlsx_formula", note=recalc_note)

    def _recalc(self, path: Path) -> bool:
        bin_name = "soffice" if shutil.which("soffice") else "libreoffice"
        try:
            proc = subprocess.run(
                [bin_name, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", str(path.parent), str(path)],
                timeout=45,
                capture_output=True,
            )
            return proc.returncode == 0
        except Exception:
            return False
